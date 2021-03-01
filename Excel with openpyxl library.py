from openpyxl import Workbook
import openpyxl


import datetime
wb = Workbook()
# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([2, 3])

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()

#Save the file
# wb.save("sample.xlsx")


book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value)
print(a3.value)
book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('sample3.xlsx')

